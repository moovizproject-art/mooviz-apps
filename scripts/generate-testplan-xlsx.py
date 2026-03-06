#!/usr/bin/env python3
"""Generate Sprint 1 Test Plan Excel from markdown source."""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_PATH = "../Docs/Testing/Testplan-sp1.md"
OUT_PATH = "../Docs/Testing/MOOVIZ-Sprint1-TestPlan.xlsx"

# Colors
BLUE = "1A73E8"
ORANGE = "FF6B35"
GREEN = "34A853"
RED = "EA4335"
LIGHT_BLUE = "E3F2FD"
LIGHT_GREEN = "E8F5E9"
LIGHT_ORANGE = "FFF3E0"
LIGHT_RED = "FFEBEE"
LIGHT_PURPLE = "F3E5F5"
LIGHT_YELLOW = "FFFDE7"
WHITE = "FFFFFF"
GRAY = "F5F5F5"

PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "P1": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "P2": PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "P3": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
}

CATEGORY_FILLS = {
    "Unit Tests": LIGHT_BLUE,
    "Integration Tests": LIGHT_GREEN,
    "UI / Component Tests": LIGHT_ORANGE,
    "End-to-End Tests": LIGHT_PURPLE,
    "Security Tests": LIGHT_RED,
    "Performance Tests": LIGHT_YELLOW,
    "Migration Tests": "E8EAF6",
    "Accessibility Tests": "F3E5F5",
}

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def parse_test_cases(md_text):
    """Parse markdown tables into structured test cases."""
    cases = []
    current_category = ""
    current_subcategory = ""

    lines = md_text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Detect category (## N. Title)
        m = re.match(r"^## \d+\.\s+(.+)$", line)
        if m:
            title = m.group(1)
            if title in CATEGORY_FILLS:
                current_category = title
            i += 1
            continue

        # Detect subcategory (### N.N Title)
        m = re.match(r"^### \d+\.\d+\s+(.+)$", line)
        if m:
            current_subcategory = m.group(1)
            i += 1
            continue

        # Detect table row with test ID
        if line.startswith("|") and not line.startswith("| ID") and not line.startswith("|--"):
            cols = [c.strip() for c in line.split("|")[1:-1]]
            if len(cols) >= 6 and re.match(r"^[A-Z]+-\d+", cols[0]):
                cases.append({
                    "category": current_category,
                    "subcategory": current_subcategory,
                    "id": cols[0],
                    "description": cols[1],
                    "preconditions": cols[2],
                    "steps": cols[3],
                    "expected": cols[4],
                    "priority": cols[5] if len(cols) > 5 else "P2",
                    "status": "",
                    "tester": "",
                    "notes": "",
                })
        i += 1

    return cases


def create_workbook(cases):
    wb = openpyxl.Workbook()

    # --- COVER SHEET ---
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.sheet_properties.tabColor = BLUE

    cover_data = [
        ("MOOVIZ Sprint 1 Test Plan", "", ""),
        ("", "", ""),
        ("Field", "Value", ""),
        ("Document", "Testplan-sp1", ""),
        ("Version", "1.0", ""),
        ("Date", "2026-03-05", ""),
        ("Author", "KAL Solutions Group", ""),
        ("Sprint", "Sprint 1 - Auth, Users, Data Models, Base Mobile", ""),
        ("Status", "Draft", ""),
        ("Total Test Cases", str(len(cases)), ""),
        ("P0 (Critical)", str(sum(1 for c in cases if c["priority"] == "P0")), ""),
        ("P1 (High)", str(sum(1 for c in cases if c["priority"] == "P1")), ""),
        ("P2 (Medium)", str(sum(1 for c in cases if c["priority"] == "P2")), ""),
    ]

    for row_idx, (a, b, c) in enumerate(cover_data, 1):
        ws_cover.cell(row=row_idx, column=1, value=a)
        ws_cover.cell(row=row_idx, column=2, value=b)
        if row_idx == 1:
            ws_cover.cell(row=1, column=1).font = Font(name="Calibri", size=20, bold=True, color=BLUE)
        elif row_idx >= 3:
            ws_cover.cell(row=row_idx, column=1).font = Font(name="Calibri", size=11, bold=True)
            ws_cover.cell(row=row_idx, column=2).font = Font(name="Calibri", size=11)

    ws_cover.column_dimensions["A"].width = 25
    ws_cover.column_dimensions["B"].width = 55

    # --- ALL TESTS SHEET ---
    ws_all = wb.create_sheet("All Test Cases")
    ws_all.sheet_properties.tabColor = GREEN

    headers = ["ID", "Category", "Subcategory", "Description", "Preconditions",
               "Steps", "Expected Result", "Priority", "Status", "Tester", "Notes"]
    col_widths = [10, 22, 28, 50, 25, 55, 50, 10, 12, 15, 25]

    header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws_all.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, case in enumerate(cases, 2):
        values = [
            case["id"], case["category"], case["subcategory"],
            case["description"], case["preconditions"], case["steps"],
            case["expected"], case["priority"], case["status"],
            case["tester"], case["notes"],
        ]
        cat_color = CATEGORY_FILLS.get(case["category"], GRAY)
        row_fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")

        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            cell.fill = row_fill

        # Priority cell gets special fill
        priority_cell = ws_all.cell(row=row_idx, column=8)
        if case["priority"] in PRIORITY_FILLS:
            priority_cell.fill = PRIORITY_FILLS[case["priority"]]
            priority_cell.alignment = Alignment(horizontal="center", vertical="top")

    # Freeze header row + auto filter
    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:K{len(cases) + 1}"

    # --- PER-CATEGORY SHEETS ---
    for cat_name, cat_color in CATEGORY_FILLS.items():
        cat_cases = [c for c in cases if c["category"] == cat_name]
        if not cat_cases:
            continue

        sheet_name = cat_name.replace("/", "-")[:31]  # Excel max 31 chars, no /
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = cat_color.replace("#", "")

        cat_headers = ["ID", "Description", "Preconditions", "Steps", "Expected Result",
                        "Priority", "Status", "Tester", "Notes"]
        cat_widths = [10, 50, 28, 55, 50, 10, 12, 15, 25]

        for col_idx, (header, width) in enumerate(zip(cat_headers, cat_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, case in enumerate(cat_cases, 2):
            values = [
                case["id"], case["description"], case["preconditions"],
                case["steps"], case["expected"], case["priority"],
                case["status"], case["tester"], case["notes"],
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

            priority_cell = ws.cell(row=row_idx, column=6)
            if case["priority"] in PRIORITY_FILLS:
                priority_cell.fill = PRIORITY_FILLS[case["priority"]]
                priority_cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cat_headers))}{len(cat_cases) + 1}"

    # --- SUMMARY SHEET ---
    ws_summary = wb.create_sheet("Summary", 1)
    ws_summary.sheet_properties.tabColor = ORANGE

    ws_summary.cell(row=1, column=1, value="Category").font = Font(bold=True, color=WHITE)
    ws_summary.cell(row=1, column=2, value="Total").font = Font(bold=True, color=WHITE)
    ws_summary.cell(row=1, column=3, value="P0").font = Font(bold=True, color=WHITE)
    ws_summary.cell(row=1, column=4, value="P1").font = Font(bold=True, color=WHITE)
    ws_summary.cell(row=1, column=5, value="P2").font = Font(bold=True, color=WHITE)
    for col in range(1, 6):
        ws_summary.cell(row=1, column=col).fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
        ws_summary.cell(row=1, column=col).border = thin_border
        ws_summary.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    row = 2
    for cat_name in CATEGORY_FILLS:
        cat_cases = [c for c in cases if c["category"] == cat_name]
        if not cat_cases:
            continue
        ws_summary.cell(row=row, column=1, value=cat_name).border = thin_border
        ws_summary.cell(row=row, column=2, value=len(cat_cases)).border = thin_border
        ws_summary.cell(row=row, column=3, value=sum(1 for c in cat_cases if c["priority"] == "P0")).border = thin_border
        ws_summary.cell(row=row, column=4, value=sum(1 for c in cat_cases if c["priority"] == "P1")).border = thin_border
        ws_summary.cell(row=row, column=5, value=sum(1 for c in cat_cases if c["priority"] == "P2")).border = thin_border

        cat_color = CATEGORY_FILLS[cat_name]
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).fill = PatternFill(start_color=cat_color, end_color=cat_color, fill_type="solid")
            ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal="center" if col > 1 else "left")
        row += 1

    # Totals row
    ws_summary.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=len(cases)).font = Font(bold=True)
    ws_summary.cell(row=row, column=3, value=sum(1 for c in cases if c["priority"] == "P0")).font = Font(bold=True)
    ws_summary.cell(row=row, column=4, value=sum(1 for c in cases if c["priority"] == "P1")).font = Font(bold=True)
    ws_summary.cell(row=row, column=5, value=sum(1 for c in cases if c["priority"] == "P2")).font = Font(bold=True)
    for col in range(1, 6):
        ws_summary.cell(row=row, column=col).border = thin_border

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 8
    ws_summary.column_dimensions["D"].width = 8
    ws_summary.column_dimensions["E"].width = 8

    return wb


def main():
    with open(MD_PATH, "r") as f:
        md_text = f.read()

    cases = parse_test_cases(md_text)
    print(f"Parsed {len(cases)} test cases")

    wb = create_workbook(cases)
    wb.save(OUT_PATH)
    print(f"Excel saved to {OUT_PATH}")


if __name__ == "__main__":
    main()
